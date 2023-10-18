
import threading
import os
import time, sys, locale
from colorama import Fore, Style

os.system("mode con cols=2000 lines=2000")
print(Fore.LIGHTGREEN_EX +
          """
                                                             _    _         ___             _____                      _____ 
                                                            | |  | |       / _ \           |_   _|                    |  _  |
                                                            | |  | | ___  / /_\ \_ __ ___    | | ___  __ _ _ __ ___    \ V / 
                                                            | |/\| |/ _ \ |  _  | '__/ _ \   | |/ _ \/ _` | '_ ` _ \   / _ \ 
                                                            \  /\  /  __/ | | | | | |  __/   | |  __/ (_| | | | | | | | |_| |
                                                             \/  \/ \___| \_| |_/_|  \___|   \_/\___|\__,_|_| |_| |_| \_____/
Package Checking.....!!!!!!!!!!!
          """)
time.sleep(5)
file_avatar = open("avatar.txt", "r").read()
for line in file_avatar.split("\n"):
    print(Fore.LIGHTBLUE_EX + line)
    time.sleep(0.02)

try:
    import PyQt5, bs4, openpyxl, requests
except:
    print("Downloading Package....... Doi Mot Chut!!!")
    time.sleep(5)
    os.system("pip install pyqt5")
    os.system("pip install requests")
    os.system("pip install openpyxl")
    os.system("exit")
from PyQt5.QtWidgets import QApplication, QMainWindow, QWidget, QFileDialog
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal
from login_ui import Ui_Form
import server1, server2
from main_gui import Ui_Form_Main
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment, PatternFill, Border, Side
import sys

print("Enjoy This Moment!!!!!!!!!!!")
time.sleep(5)

class MyWindow(QMainWindow):
    status_update = pyqtSignal(int, str)
    violated = pyqtSignal(int)
    not_violated = pyqtSignal(int)

    def __init__(self):
        super().__init__()
        self.ui = Ui_Form()
        self.ui.setupUi(self)
        self.main_window = QtWidgets.QWidget()
        self.ui_main = Ui_Form_Main()
        self.ui_main.setupUi(self.main_window)
        self.ui.pushButton_5.clicked.connect(self.hide_window)
        self.ui.pushButton_2.clicked.connect(self.close_window)
        self.ui.pushButton.clicked.connect(self.open_main)
        self.ui_main.pushButton.clicked.connect(self.open_file)
        self.ui_main.pushButton_2.clicked.connect(self.sever_check_thread)
        self.ui_main.pushButton_3.clicked.connect(self.start_check)
        self.ui_main.pushButton_5.clicked.connect(self.clear_table)
        self.ui_main.pushButton_6.clicked.connect(self.extract_file)
        self.data_licensekey = []
        self.data_violated = []
        self.data_violated_extract = []
        self.status_update.connect(self.update_status)
        self.violated.connect(self.violated_update)
        self.not_violated.connect(self.not_violated_update)
        self.violated_count = 0
        self.not_violated_count = 0
        self.stop = False

        self.user = "NONE"

    def hide_window(self):
        self.showMinimized()

    def close_window(self):
        self.close()

    def open_main(self):
        self.main_window.show()
        self.user = self.ui.lineEdit.text()
        self.ui_main.label_21.setText(self.user)
        if self.user == "":
            self.ui_main.label_21.setText("Guest")
        self.close()

    def open_file(self):
        if self.ui_main.radioButton.isChecked():
            options = QFileDialog.Options()
            options |= QFileDialog.ReadOnly
            file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Textfile (*.txt)",
                                                       options=options)
            print(file_name)
            if file_name:
                with open(file_name, 'r') as file:
                    file_content = file.read().split("\n")
                    self.ui_main.label_40.setText(str(len(file_content)))
                    for line in file_content:
                        self.data_licensekey.append(line.replace(".", "").replace("-", "").upper())
            self.ui_main.tableWidget.setRowCount(len(self.data_licensekey))
            for index in range(len(self.data_licensekey)):
                self.ui_main.tableWidget.setItem(index, 0, QtWidgets.QTableWidgetItem(self.user))
                self.ui_main.tableWidget.setItem(index, 1, QtWidgets.QTableWidgetItem(self.data_licensekey[index]))
                self.ui_main.tableWidget.setItem(index, 2, QtWidgets.QTableWidgetItem(str(datetime.date.today())))
                self.ui_main.tableWidget.setItem(index, 3, QtWidgets.QTableWidgetItem("Not Check!"))
        if self.ui_main.radioButton_3.isChecked():
            options = QFileDialog.Options()
            options |= QFileDialog.ReadOnly
            file_name, _ = QFileDialog.getOpenFileName(self, "Open File", "", "Excel (*.xlsx)",
                                                       options=options)
            if file_name:
                excel_file_path = str(file_name)

                desired_column_index = int(self.ui_main.lineEdit.text())
                if desired_column_index >= 1:
                    desired_column_index -= 1
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet = workbook.active
                column_data = []
                for row in sheet.iter_rows(min_col=desired_column_index + 1, max_col=desired_column_index + 1):
                    for cell in row:
                        self.data_licensekey.append(str(cell.value).replace(".", "").replace("-", "").upper())
                # Now, you can work with the data in the column_data list
                self.data_licensekey.pop(0)
                self.ui_main.tableWidget.setRowCount(len(self.data_licensekey))
                for index in range(len(self.data_licensekey)):
                    self.ui_main.tableWidget.setItem(index, 0, QtWidgets.QTableWidgetItem(self.user))
                    self.ui_main.tableWidget.setItem(index, 1, QtWidgets.QTableWidgetItem(self.data_licensekey[index]))
                    self.ui_main.tableWidget.setItem(index, 2, QtWidgets.QTableWidgetItem(str(datetime.date.today())))
                    self.ui_main.tableWidget.setItem(index, 3, QtWidgets.QTableWidgetItem("Not Check!"))

    def sever_check(self):
        self.ui_main.label_32.setStyleSheet("color:#FFBF00")
        self.ui_main.label_32.setText("Checking....")
        if self.ui_main.radioButton_4.isChecked():
            result = server1.check_status()
            if result == "Active":
                self.ui_main.label_32.setStyleSheet("color:#008000")
                self.ui_main.label_32.setText("Active")
            else:
                self.ui_main.label_32.setText("Error")

        elif self.ui_main.radioButton_5.isChecked():
            result = server2.check_status()
            if result == "Active":
                self.ui_main.label_32.setStyleSheet("color:#008000")
                self.ui_main.label_32.setText("Active")
            else:
                self.ui_main.label_32.setText("Error")
        else:
            self.ui_main.label_32.setText("Server not selected")
        if self.stop == True:
            self.ui_main.label_32.setText("Stopped")

    def sever_check_thread(self):
        threading.Thread(target=self.sever_check).start()

    def start_check(self):
        for i in range(len(self.data_licensekey)):
            threading.Thread(target=self.check_main, args=(i, self.data_licensekey[i]), ).start()

    def stop_button(self):
        self.stop = True

    def check_main(self, index, license_key):
        self.status_update.emit(index, "Checking...🚀🚀🚀")
        if self.ui_main.radioButton_4.isChecked():
            result = server1.getdata(license_key)
            if result == "NOT VIOLATED":
                self.not_violated.emit(1)
                self.status_update.emit(index, "NOT VIOLATED! ")
            elif result == "ERROR":
                self.status_update.emit(index, "ERROR!🛑🛑🛑")
            else:
                self.violated.emit(1)
                self.status_update.emit(index, "VIOLATED!🛑🛑🛑")
                self.data_violated.append(result)
        elif self.ui_main.radioButton_5.isChecked():
            #result = server2.getdata(license_key)
          '''
            if result == "NOT VIOLATED":
                self.not_violated.emit(1)
                self.status_update.emit(index, "NOT VIOLATED! ")
            elif result == "ERROR":
                self.status_update.emit(index, "ERROR!🛑🛑🛑")
            else:
                self.violated.emit(1)
                self.status_update.emit(index, "VIOLATED!🛑🛑🛑")
                self.data_violated.append(result)
          '''
          self.status_update.emit(index, "Server Bao Tri ^^^^")
        else:
            self.status_update.emit(index, "Please Select A Server🛑🛑🛑")

        if self.stop == True:
            self.ui_main.label_32.setText("Stopped")

    def update_status(self, index, text):
        self.ui_main.tableWidget.setItem(index, 3, QtWidgets.QTableWidgetItem(text))

    def violated_update(self, num):
        self.violated_count += num
        self.ui_main.label_45.setText(str(self.violated_count))

    def not_violated_update(self, num):
        self.not_violated_count += num
        self.ui_main.label_44.setText(str(self.not_violated_count))

    def clear_table(self):
        self.ui_main.tableWidget.setRowCount(0)

    def extract_file(self):
        for i in range(len(self.data_violated)):
            data_tuple = (i + 1,)
            for key,value in self.data_violated[i].items():
                for j in range(len(value)):
                    value[j] = "-" + value[j]
                data_tuple += ("\n".join(value),)
            data_tuple += ("-" + str(datetime.date.today()),)
            self.data_violated_extract.append(data_tuple)

        # Tạo một tệp Excel mới
        workbook = openpyxl.Workbook()

        # Chọn trang tính mặc định (Sheet 1)
        sheet = workbook.active

        # Đặt tiêu đề cho các cột
        sheet['A1'] = 'STT'
        sheet['B1'] = 'Biển Số Xe'
        sheet['C1'] = 'Số lỗi vi phạm'
        sheet['D1'] = 'Màu biển'
        sheet['E1'] = 'Loại phương tiện'
        sheet['F1'] = 'Lỗi vi phạm'
        sheet['G1'] = 'Địa điểm vi phạm'
        sheet['H1'] = 'Thời gian vi phạm'
        sheet['I1'] = 'Trạng thái'
        sheet['J1'] = 'Đơn vị phát hiện vi phạm'
        sheet['K1'] = 'Nơi giải quyết vụ việc'
        sheet['L1'] = 'Thời gian Kiểm Tra'
        sheet.column_dimensions['A'].width = 5
        sheet.column_dimensions['B'].width = 13
        sheet.column_dimensions['C'].width = 7
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 50
        sheet.column_dimensions['G'].width = 50
        sheet.column_dimensions['H'].width = 50
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 50
        sheet.column_dimensions['K'].width = 50
        sheet.column_dimensions['L'].width = 15
        sheet.row_dimensions[1].height = 50
        list_head = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1']
        # Mẫu dữ liệu để thêm vào tệp Excel

        # Need modify
        data = self.data_violated_extract

        fill = PatternFill(start_color='40e0d0', end_color='40e0d0', fill_type='solid')
        for name_row in list_head:
            sheet[f'{name_row}'].fill = fill

        # Thêm dữ liệu vào tệp Excel
        for row in data:
            sheet.append(row)

        border = Border(left=Side(border_style='thin'),
                        right=Side(border_style='thin'),
                        top=Side(border_style='thin'),
                        bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = border

        # Đặt cỡ chữ cho các ô trong cột "Lỗi Vi Phạm"
        for row in sheet.iter_rows(min_row=1, max_row=1, min_col=0, max_col=12):
            for cell in row:
                cell.font = Font(size=11)  # Đặt cỡ chữ là 12
                cell.alignment = Alignment(wrapText=True, horizontal='center', vertical='center')

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=0, max_col=12):
            for cell in row:
                cell.font = Font(size=11)  # Đặt cỡ chữ là 12
                cell.alignment = Alignment(wrapText=True, horizontal='left', vertical='center')
        # Lưu tệp Excel
        workbook.save('danh_sach_vi_pham.xlsx')

        # Đóng tệp Excel
        workbook.close()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MyWindow()
    window.show()
    sys.exit(app.exec_())
